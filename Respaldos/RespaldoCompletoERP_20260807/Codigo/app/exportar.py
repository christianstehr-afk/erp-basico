"""
Módulo 5 · Generación de archivos para el export contable.

- construir_excel: listado de movimientos (ingresos/egresos) a un .xlsx.
- construir_zip_rendiciones: un PDF por rendición (información + adjuntos, una
  imagen por página) empaquetados en un .zip.
- parsear_cartola_banco_chile / comparar_cc / construir_excel_comparacion:
  comparación de la cartola del banco contra los movimientos de la app.

Estas funciones no tocan la base de datos; reciben los datos ya consultados.
"""
from __future__ import annotations

import io
import os
import re
import unicodedata
import zipfile
from pathlib import Path

from .db import codigo_rendicion

# Extensiones de imagen que se pintan "una por página" en el PDF.
IMG_EXT = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp"}

# Logo E-Auto para la portada del PDF de rendición. Se descarga UNA vez y se
# cachea en disco (evita depender de internet en cada PDF); si la descarga
# falla, la portada se dibuja igual pero sin logo (nunca debe tumbar el PDF).
_LOGO_URL = "https://www.e-auto.global/assets/logos/eauto_logo_new.webp"
_LOGO_CACHE = Path(os.environ.get(
    "LOGO_CACHE_PATH", Path(__file__).resolve().parent.parent / "data" / "eauto_logo.png"
))


def _logo_reader():
    """Devuelve un ImageReader del logo E-Auto (cacheado), o None si no se pudo obtener."""
    from reportlab.lib.utils import ImageReader

    try:
        if not (_LOGO_CACHE.exists() and _LOGO_CACHE.stat().st_size > 0):
            import requests
            from PIL import Image

            resp = requests.get(_LOGO_URL, timeout=10)
            resp.raise_for_status()
            img = Image.open(io.BytesIO(resp.content)).convert("RGBA")
            _LOGO_CACHE.parent.mkdir(parents=True, exist_ok=True)
            img.save(_LOGO_CACHE, "PNG")
        return ImageReader(str(_LOGO_CACHE))
    except Exception:
        return None


def _miles(n: int) -> str:
    """Formato de miles con punto (estilo chileno): 1234567 -> 1.234.567."""
    return "{:,.0f}".format(n or 0).replace(",", ".")


def _nombre_archivo(nombre: str, sufijo: str) -> str:
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", (nombre or "").strip()) or "rendicion"
    return f"{base[:80]}{sufijo}"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def construir_excel(movimientos: list[dict], desde: str, hasta: str) -> bytes:
    """Arma el .xlsx del listado de movimientos y lo devuelve en bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    verde = "FF009406"
    tinta = "FF0A0A0A"
    rojo = "FFC0392B"
    gris = "FFF2F2F2"

    # Título y rango.
    ws["A1"] = "Movimientos de caja"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=tinta)
    ws["A2"] = f"Rango: {desde} a {hasta}"
    ws["A2"].font = Font(name="Calibri", size=10, color="FF666666")

    encabezados = ["Fecha", "Ingreso/Egreso", "Descripción", "Centro", "Monto"]
    fila_head = 4
    borde = Border(bottom=Side(style="thin", color="FFDDDDDD"))
    for col, texto in enumerate(encabezados, start=1):
        c = ws.cell(row=fila_head, column=col, value=texto)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=tinta)
        c.alignment = Alignment(horizontal="right" if texto == "Monto" else "left")

    total_ing = 0
    total_egr = 0
    fila = fila_head + 1
    for m in movimientos:
        es_ing = m["flujo"] == "Ingreso"
        if es_ing:
            total_ing += m["monto"]
        else:
            total_egr += m["monto"]
        ws.cell(row=fila, column=1, value=m["fecha"])
        cflujo = ws.cell(row=fila, column=2, value=m["flujo"])
        cflujo.font = Font(bold=True, color=verde if es_ing else rojo)
        ws.cell(row=fila, column=3, value=m["descripcion"])
        ws.cell(row=fila, column=4, value=m.get("centro") or "")
        cmonto = ws.cell(row=fila, column=5, value=m["monto"])
        cmonto.number_format = '"$"#,##0'
        cmonto.alignment = Alignment(horizontal="right")
        for col in range(1, 6):
            ws.cell(row=fila, column=col).border = borde
        fila += 1

    # Totales.
    fila += 1
    ws.cell(row=fila, column=4, value="Total ingresos").font = Font(bold=True, color=verde)
    ci = ws.cell(row=fila, column=5, value=total_ing)
    ci.number_format = '"$"#,##0'; ci.font = Font(bold=True, color=verde)
    fila += 1
    ws.cell(row=fila, column=4, value="Total egresos").font = Font(bold=True, color=rojo)
    ce = ws.cell(row=fila, column=5, value=total_egr)
    ce.number_format = '"$"#,##0'; ce.font = Font(bold=True, color=rojo)
    fila += 1
    ws.cell(row=fila, column=4, value="Neto (ingresos − egresos)").font = Font(bold=True, color=tinta)
    cn = ws.cell(row=fila, column=5, value=total_ing - total_egr)
    cn.number_format = '"$"#,##0'; cn.font = Font(bold=True, color=tinta)
    ws.cell(row=fila, column=4).fill = PatternFill("solid", fgColor=gris)
    ws.cell(row=fila, column=5).fill = PatternFill("solid", fgColor=gris)

    anchos = {1: 14, 2: 16, 3: 55, 4: 14, 5: 16}
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"

    # --- Hoja 2: resumen por centro de costo/ingreso ---------------------
    # Agrupa los movimientos por su centro ("LINEA-CAT"). Un pago de rendición
    # con ítems de varios centros aparece con el rótulo combinado ("A / B");
    # lo no imputado suma en "(sin imputar)".
    ws2 = wb.create_sheet("Por centro")
    ws2["A1"] = "Resumen por centro de costo/ingreso"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True, color=tinta)
    ws2["A2"] = f"Rango: {desde} a {hasta}"
    ws2["A2"].font = Font(name="Calibri", size=10, color="FF666666")
    for col, texto in enumerate(["Centro", "Ingresos", "Egresos", "Neto"], start=1):
        c = ws2.cell(row=4, column=col, value=texto)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=tinta)
        if col > 1:
            c.alignment = Alignment(horizontal="right")
    # Cuando el movimiento viene de una factura/rendición repartida en varios
    # centros ("centros_detalle": lista de (centro, monto) que suma el monto
    # del movimiento — ver db.movimientos_en_rango), cada centro aporta solo
    # su parte proporcional, no el monto completo. Así el TAG de carreteras
    # pagado 60% Gecko / 40% flota queda bien repartido en este resumen.
    resumen: dict = {}
    for m in movimientos:
        detalle = m.get("centros_detalle") or [(m.get("centro") or "(sin imputar)", m["monto"])]
        for clave, monto_parte in detalle:
            ing, egr = resumen.get(clave, (0, 0))
            if m["flujo"] == "Ingreso":
                ing += monto_parte
            else:
                egr += monto_parte
            resumen[clave] = (ing, egr)
    fila2 = 5
    for clave in sorted(resumen, key=lambda k: (k == "(sin imputar)", k)):
        ing, egr = resumen[clave]
        ws2.cell(row=fila2, column=1, value=clave)
        for col, valor in ((2, ing), (3, egr), (4, ing - egr)):
            c = ws2.cell(row=fila2, column=col, value=valor)
            c.number_format = '"$"#,##0'
            c.alignment = Alignment(horizontal="right")
        for col in range(1, 5):
            ws2.cell(row=fila2, column=col).border = borde
        fila2 += 1
    for col, w in {1: 26, 2: 16, 3: 16, 4: 16}.items():
        ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def construir_excel_logs(logs: list) -> bytes:
    """Arma el .xlsx del log de auditoría (fecha, hora, acción, usuario) y lo
    devuelve en bytes. `logs` es el resultado de db.listar_logs (más recientes
    primero)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Log"

    tinta = "FF0A0A0A"

    ws["A1"] = "Log de operaciones · ERP e-auto"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=tinta)
    ws["A2"] = f"Total de registros: {len(logs)}"
    ws["A2"].font = Font(name="Calibri", size=10, color="FF666666")

    encabezados = ["Fecha", "Hora", "Acción", "Usuario"]
    fila_head = 4
    borde = Border(bottom=Side(style="thin", color="FFDDDDDD"))
    for col, texto in enumerate(encabezados, start=1):
        c = ws.cell(row=fila_head, column=col, value=texto)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=tinta)
        c.alignment = Alignment(horizontal="left")

    fila = fila_head + 1
    for log in logs:
        ws.cell(row=fila, column=1, value=log["fecha"])
        ws.cell(row=fila, column=2, value=log["hora"])
        ws.cell(row=fila, column=3, value=log["accion"])
        ws.cell(row=fila, column=4, value=log["usuario"] or "")
        for col in range(1, 5):
            ws.cell(row=fila, column=col).border = borde
        fila += 1

    anchos = {1: 13, 2: 11, 3: 70, 4: 16}
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF por rendición
# ---------------------------------------------------------------------------

def _pagina_info(c, r, items, pagos) -> None:
    """Dibuja la página de información (portada) de una rendición."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm

    W, H = A4
    mx = 20 * mm
    y = H - 22 * mm

    # Logo: justificado a la derecha, a la misma altura del título (nombre),
    # sin ocupar espacio en el flujo vertical del texto de la izquierda.
    logo = _logo_reader()
    if logo is not None:
        try:
            iw, ih = logo.getSize()
            logo_h = 14 * mm
            logo_w = logo_h * iw / ih
            titulo_y = y - 9 * mm  # misma y donde se dibuja el título más abajo
            c.drawImage(logo, W - mx - logo_w, titulo_y - 3 * mm, logo_w, logo_h,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass  # un logo que no se pudo dibujar no debe tumbar el PDF

    c.setFillColorRGB(0, 0.58, 0.023)  # verde e-auto
    c.setFont("Helvetica-Bold", 9)
    c.drawString(mx, y, f"E-AUTO · RENDICIÓN {codigo_rendicion(r['id'])}")
    y -= 9 * mm
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(mx, y, (r["nombre"] or "")[:70])
    y -= 7 * mm
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.setFont("Helvetica", 10)
    c.drawString(mx, y, f"Fecha de rendición: {r['fecha']}")
    y -= 12 * mm

    total = r["total"] or 0
    pagado = r["pagado"] or 0
    saldo = total - pagado
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(mx, y, f"Total: ${_miles(total)}    Pagado: ${_miles(pagado)}    Saldo: ${_miles(saldo)}")
    y -= 12 * mm

    # Ítems
    c.setFont("Helvetica-Bold", 10)
    c.drawString(mx, y, "Ítems")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, y, "DESCRIPCIÓN")
    c.drawString(mx + 95 * mm, y, "N° DOC")
    c.drawRightString(W - mx, y, "MONTO")
    y -= 2 * mm
    c.setStrokeColorRGB(0.85, 0.85, 0.85)
    c.line(mx, y, W - mx, y)
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    for it in items:
        if y < 30 * mm:
            c.showPage()
            y = H - 22 * mm
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.1, 0.1, 0.1)
        c.drawString(mx, y, (it["descripcion"] or "")[:60])
        c.drawString(mx + 95 * mm, y, str(it["numero_doc"] or "—")[:18])
        c.drawRightString(W - mx, y, f"${_miles(it['monto'])}")
        y -= 5.5 * mm

    # Pagos
    y -= 6 * mm
    if y < 40 * mm:
        c.showPage()
        y = H - 22 * mm
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.drawString(mx, y, "Pagos registrados")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, y, "FECHA")
    c.drawRightString(W - mx, y, "MONTO")
    y -= 2 * mm
    c.line(mx, y, W - mx, y)
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    for p in pagos:
        if y < 20 * mm:
            c.showPage()
            y = H - 22 * mm
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.1, 0.1, 0.1)
        c.drawString(mx, y, str(p["fecha"]))
        c.drawRightString(W - mx, y, f"${_miles(p['monto'])}")
        y -= 5.5 * mm
    c.showPage()


def _pagina_imagen(c, path: str, caption: str) -> None:
    """Dibuja una imagen ocupando la página (con leyenda)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader

    W, H = A4
    mx = 15 * mm
    top = H - 15 * mm
    img = ImageReader(path)
    iw, ih = img.getSize()
    max_w = W - 2 * mx
    max_h = H - 30 * mm  # deja espacio para la leyenda
    escala = min(max_w / iw, max_h / ih)
    w = iw * escala
    h = ih * escala
    x = (W - w) / 2
    y = top - h
    c.drawImage(img, x, y, w, h, preserveAspectRatio=True, anchor="n")
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, 10 * mm, f"Adjunto: {caption}"[:110])
    c.showPage()


def _pdf_de_rendicion(r, items, adjuntos, pagos) -> bytes:
    """Devuelve el PDF (bytes) de una rendición: info + imágenes + PDFs adjuntos."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from pypdf import PdfReader, PdfWriter

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _pagina_info(c, r, items, pagos)
    # Imágenes: una por página.
    for a in adjuntos:
        ext = Path(a["nombre_archivo"] or "").suffix.lower()
        if ext in IMG_EXT and Path(a["path"]).exists():
            try:
                _pagina_imagen(c, a["path"], a["nombre_archivo"])
            except Exception:
                pass
    c.save()

    writer = PdfWriter()
    for page in PdfReader(io.BytesIO(buf.getvalue())).pages:
        writer.add_page(page)
    # PDFs adjuntos: se anexan al final tal cual.
    for a in adjuntos:
        ext = Path(a["nombre_archivo"] or "").suffix.lower()
        if ext == ".pdf" and Path(a["path"]).exists():
            try:
                for page in PdfReader(a["path"]).pages:
                    writer.add_page(page)
            except Exception:
                pass
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _pdf_de_gestion_pago(f, pagos, cfg, incluye_original: bool = False,
                         incluye_adjuntos: bool = False) -> bytes:
    """Devuelve el PDF (bytes) del detalle de gestión de una factura (pago a
    proveedores / cobro a clientes): resumen + movimientos parciales.

    Si `incluye_original` es True, deja una nota indicando que el documento
    original de la factura viene a continuación (el llamador es quien anexa
    esas páginas, con `anexar_pdf`, porque requiere una sesión SII activa).
    Si `incluye_adjuntos` es True, deja una nota equivalente para los
    adjuntos de la gestión (el llamador los anexa con `anexar_archivos`)."""
    import textwrap

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    from .db import codigo_rendicion

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    mx = 20 * mm
    y = H - 22 * mm

    logo = _logo_reader()
    if logo is not None:
        try:
            iw, ih = logo.getSize()
            logo_h = 14 * mm
            logo_w = logo_h * iw / ih
            titulo_y = y - 9 * mm
            c.drawImage(logo, W - mx - logo_w, titulo_y - 3 * mm, logo_w, logo_h,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    rechazada = bool(f["fecha_reclamo"])
    total = f["total"] or 0
    pagado = f["pagado"] or 0
    saldo = total - pagado
    pagada = pagado >= total
    estado = "Rechazada" if rechazada else (cfg["estado_ok"] if pagada else "Pendiente")

    c.setFillColorRGB(0, 0.58, 0.023)  # verde e-auto
    c.setFont("Helvetica-Bold", 9)
    c.drawString(mx, y, f"E-AUTO · {cfg['titulo'].upper()}")
    y -= 9 * mm
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(mx, y, f"{f['documento']} · Folio {f['folio']}"[:80])
    y -= 7 * mm
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.setFont("Helvetica", 10)
    c.drawString(mx, y, f"{f['razon_social'] or ''} · {f['rut_contraparte'] or ''}"[:100])
    y -= 6 * mm
    c.drawString(mx, y, f"Fecha de emisión: {f['fecha_emision'] or '—'}")
    y -= 12 * mm

    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(
        mx, y,
        f"Monto factura: ${_miles(total)}    {cfg['label_pagado']}: ${_miles(pagado)}    "
        f"Saldo: ${_miles(saldo)}",
    )
    y -= 7 * mm
    c.setFont("Helvetica-Bold", 10)
    if rechazada:
        c.setFillColorRGB(0.75, 0.15, 0.15)
    elif pagada:
        c.setFillColorRGB(0, 0.58, 0.023)
    else:
        c.setFillColorRGB(0.05, 0.4, 0.85)
    c.drawString(mx, y, f"Estado: {estado}")
    y -= 10 * mm

    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(mx, y, f"Fecha de {cfg['accion']} tope: {f['fecha_pago_tope'] or '—'}")
    y -= 10 * mm

    if f["descripcion"]:
        c.setFont("Helvetica-Bold", 10)
        c.drawString(mx, y, "Descripción")
        y -= 6 * mm
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.15, 0.15, 0.15)
        for linea in textwrap.wrap(f["descripcion"], 105) or [""]:
            if y < 25 * mm:
                c.showPage()
                y = H - 22 * mm
                c.setFont("Helvetica", 9)
                c.setFillColorRGB(0.15, 0.15, 0.15)
            c.drawString(mx, y, linea)
            y -= 5 * mm
        y -= 6 * mm

    # Movimientos parciales
    if y < 40 * mm:
        c.showPage()
        y = H - 22 * mm
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(mx, y, f"{cfg['accion'].capitalize()}s parciales")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, y, "FECHA")
    c.drawString(mx + 40 * mm, y, "VÍA")
    c.drawRightString(W - mx, y, "MONTO")
    y -= 2 * mm
    c.setStrokeColorRGB(0.85, 0.85, 0.85)
    c.line(mx, y, W - mx, y)
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    if pagos:
        for p in pagos:
            if y < 20 * mm:
                c.showPage()
                y = H - 22 * mm
                c.setFont("Helvetica", 9)
                c.setFillColorRGB(0.1, 0.1, 0.1)
            if p["rendicion_id"]:
                via = f"Rendición {codigo_rendicion(p['rendicion_id'])}"
            elif p["externo"]:
                via = "Externo"
            else:
                via = "—"
            c.drawString(mx, y, str(p["fecha"]))
            c.drawString(mx + 40 * mm, y, via[:38])
            c.drawRightString(W - mx, y, f"${_miles(p['monto'])}")
            y -= 5.5 * mm
    else:
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.drawString(mx, y, f"Aún no hay {cfg['accion']}s registrados.")
        y -= 5.5 * mm

    if incluye_original:
        y -= 8 * mm
        if y < 20 * mm:
            c.showPage()
            y = H - 22 * mm
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(mx, y, "— Documento original de la factura a continuación —")

    if incluye_adjuntos:
        y -= 6 * mm
        if y < 20 * mm:
            c.showPage()
            y = H - 22 * mm
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(mx, y, "— Adjuntos de la gestión a continuación —")

    c.showPage()
    c.save()
    return buf.getvalue()


def anexar_archivos(base: bytes, adjuntos: list[dict]) -> bytes:
    """Anexa al final de un PDF los archivos adjuntos de la gestión de una
    factura (ver `factura_adjuntos` en db.py): imágenes se dibujan una por
    página (con leyenda con el nombre del archivo), PDFs se agregan tal
    cual. Un adjunto que no existe en disco o no se puede leer se salta sin
    interrumpir a los demás — nunca debe tumbar la generación del PDF.
    """
    if not adjuntos:
        return base

    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    resultado = base

    imagenes = [
        a for a in adjuntos
        if Path(a["nombre_archivo"] or "").suffix.lower() in IMG_EXT
        and a["path"] and Path(a["path"]).exists()
    ]
    if imagenes:
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        for a in imagenes:
            try:
                _pagina_imagen(c, a["path"], a["nombre_archivo"])
            except Exception:
                pass  # una imagen que no se pudo dibujar no debe tumbar el resto
        c.save()
        resultado = anexar_pdf(resultado, buf.getvalue())

    for a in adjuntos:
        ext = Path(a["nombre_archivo"] or "").suffix.lower()
        if ext == ".pdf" and a["path"] and Path(a["path"]).exists():
            try:
                resultado = anexar_pdf(resultado, Path(a["path"]).read_bytes())
            except Exception:
                pass  # un PDF adjunto ilegible no debe tumbar el resto

    return resultado


def anexar_pdf(base: bytes, extra: bytes | None) -> bytes:
    """Agrega las páginas de `extra` al final de `base`.

    Si `extra` es None o no se puede leer como PDF, devuelve `base` sin
    cambios (nunca debe tumbar la generación del PDF de gestión).
    """
    if not extra:
        return base
    from pypdf import PdfReader, PdfWriter

    try:
        writer = PdfWriter()
        for page in PdfReader(io.BytesIO(base)).pages:
            writer.add_page(page)
        for page in PdfReader(io.BytesIO(extra)).pages:
            writer.add_page(page)
        out = io.BytesIO()
        writer.write(out)
        return out.getvalue()
    except Exception:
        return base


def construir_zip_rendiciones(rendiciones: list[dict]) -> bytes:
    """Empaqueta un PDF por rendición en un .zip.

    `rendiciones` es una lista de dicts con: r (fila), items, adjuntos, pagos.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        usados: set[str] = set()
        for rd in rendiciones:
            r = rd["r"]
            pdf = _pdf_de_rendicion(r, rd["items"], rd["adjuntos"], rd["pagos"])
            cod = codigo_rendicion(r["id"])
            nombre = f"{cod}_" + _nombre_archivo(r["nombre"], ".pdf")
            while nombre in usados:
                nombre = f"{cod}_" + _nombre_archivo(r["nombre"], "_x.pdf")
            usados.add(nombre)
            zf.writestr(nombre, pdf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Módulo 5 · PDF del listado de Movimientos CC
# ---------------------------------------------------------------------------

_ORIGEN_LABEL = {"factura": "Factura", "rendicion": "Rendición", "manual": "Manual"}


def construir_pdf_movimientos_cc(movs: list, desde: str, hasta: str,
                                 total_ingresos: int, total_egresos: int) -> bytes:
    """PDF del listado de Movimientos CC, en el mismo orden y rango con que se
    armó `movs` en pantalla (ver GET /movimientos/pdf en main.py: usa
    exactamente los mismos `movs`/`desde`/`hasta` que /movimientos)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    mx = 15 * mm
    col_fecha = mx
    col_tipo = mx + 22 * mm
    col_desc = mx + 42 * mm
    col_monto_r = W - mx - 33 * mm  # borde derecho de la columna Monto
    col_origen = W - mx - 27 * mm

    def _encabezado(y: float) -> float:
        logo = _logo_reader()
        if logo is not None:
            try:
                iw, ih = logo.getSize()
                logo_h = 12 * mm
                logo_w = logo_h * iw / ih
                c.drawImage(logo, W - mx - logo_w, y - 9 * mm, logo_w, logo_h,
                            preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        c.setFillColorRGB(0, 0.58, 0.023)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(mx, y, "E-AUTO · MOVIMIENTOS CC")
        y -= 9 * mm
        c.setFillColorRGB(0.04, 0.04, 0.04)
        c.setFont("Helvetica-Bold", 17)
        c.drawString(mx, y, "Movimientos CC")
        y -= 7 * mm
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.setFont("Helvetica", 10)
        c.drawString(mx, y, f"Del {desde} al {hasta}")
        y -= 8 * mm
        neto = (total_ingresos or 0) - (total_egresos or 0)
        c.setFillColorRGB(0.04, 0.04, 0.04)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(
            mx, y,
            f"Ingresos: ${_miles(total_ingresos)}    Egresos: ${_miles(total_egresos)}    Neto: ${_miles(neto)}",
        )
        y -= 10 * mm
        return _encabezado_tabla(y)

    def _encabezado_tabla(y: float) -> float:
        c.setFont("Helvetica-Bold", 8)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(col_fecha, y, "FECHA")
        c.drawString(col_tipo, y, "TIPO")
        c.drawString(col_desc, y, "DESCRIPCIÓN")
        c.drawRightString(col_monto_r, y, "MONTO")
        c.drawString(col_origen, y, "ORIGEN")
        y -= 2 * mm
        c.setStrokeColorRGB(0.85, 0.85, 0.85)
        c.line(mx, y, W - mx, y)
        y -= 5.5 * mm
        c.setFont("Helvetica", 8.5)
        c.setFillColorRGB(0.1, 0.1, 0.1)
        return y

    y = _encabezado(H - 22 * mm)

    for m in movs:
        if y < 20 * mm:
            c.showPage()
            y = _encabezado_tabla(H - 20 * mm)
        ingreso = m["flujo"] == "Ingreso"
        color_flujo = (0, 0.45, 0.1) if ingreso else (0.231, 0.510, 0.965)

        c.setFillColorRGB(*color_flujo)
        c.setFont("Helvetica", 8.5)
        c.drawString(col_fecha, y, str(m["fecha"] or ""))
        c.drawString(col_tipo, y, m["flujo"] or "")

        c.setFillColorRGB(0.1, 0.1, 0.1)
        c.drawString(col_desc, y, (m["descripcion"] or "")[:48])

        c.setFillColorRGB(*color_flujo)
        c.drawRightString(col_monto_r, y, f"${_miles(m['monto'])}")

        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.drawString(col_origen, y, _ORIGEN_LABEL.get(m["origen"], m["origen"] or ""))

        y -= 5.5 * mm

    if not movs:
        c.setFont("Helvetica", 10)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(mx, y, "No hay movimientos registrados en este rango.")

    c.showPage()
    c.save()
    return buf.getvalue()


# Carpetas/archivos que nunca deben ir en el respaldo de código, aunque
# aparezcan bajo code_dir en un entorno local (en Railway ni siquiera existen,
# ver .dockerignore): son datos, entornos o carpetas de trabajo de Christian,
# no código de la app.
_EXCLUIR_RESPALDO_CODIGO = {
    "data", ".venv", "venv", "__pycache__", ".git", "Respaldos", "Log",
    "temp", "node_modules", ".DS_Store",
}


def _leeme_respaldo(fecha: str) -> str:
    return (
        "ERP Básico · Respaldo completo\n"
        "===============================\n\n"
        f"Generado: {fecha}\n\n"
        "Qué trae este .zip:\n"
        "- BaseDatos/erp.db      Toda la información: facturas, pagos, rendiciones,\n"
        "                        movimientos CC y el log de auditoría.\n"
        "- Adjuntos/Rendiciones  Boletas/facturas subidas a cada rendición.\n"
        "- Adjuntos/Facturas     Documentos subidos en la gestión de pago/cobro de\n"
        "                        cada factura (ojo: los PDF de las facturas del SII\n"
        "                        NO se guardan acá, se piden al SII al momento de\n"
        "                        verlos con sesión activa; no hace falta respaldarlos).\n"
        "- Codigo/               Copia completa del código de la app, tal como estaba\n"
        "                        corriendo al momento de generar este respaldo.\n\n"
        "Cómo reconstruir todo desde cero frente a un desastre:\n"
        "1. Subir la carpeta Codigo/ a un repositorio de GitHub nuevo (o al mismo de\n"
        "   siempre) y desplegarlo igual que la primera vez (ver Codigo/DEPLOY.md).\n"
        "2. En el servidor nuevo, copiar BaseDatos/erp.db a la ruta de la variable de\n"
        "   entorno DB_PATH (en Railway: el volumen persistente, por defecto /data/erp.db).\n"
        "3. Copiar Adjuntos/Rendiciones a la ruta de ADJUNTOS_DIR (por defecto\n"
        "   /data/adjuntos/rendiciones) y Adjuntos/Facturas a la ruta de\n"
        "   ADJUNTOS_FACTURAS_DIR (por defecto /data/adjuntos/facturas).\n"
        "4. Iniciar la app normalmente: al arrancar crea/actualiza el esquema de la\n"
        "   base de datos solo, sin tocar los datos ya copiados.\n"
    )


def construir_respaldo_completo(db_bytes: bytes, adjuntos_dir: Path,
                                adjuntos_facturas_dir: Path, code_dir: Path,
                                fecha: str) -> bytes:
    """Arma un .zip con TODO lo necesario para reconstruir la app desde cero
    frente a un desastre informático: la base de datos, los adjuntos
    (rendiciones y gestión de pago de facturas) y una copia del código fuente
    tal como está corriendo ahora mismo. Ver GET /respaldo en main.py (botón
    "Descargar Respaldo" del Cockpit)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("BaseDatos/erp.db", db_bytes)

        def _agregar_carpeta(origen: Path, prefijo: str) -> None:
            if not origen.exists():
                return
            for p in origen.rglob("*"):
                if p.is_file():
                    zf.write(p, f"{prefijo}/{p.relative_to(origen)}")

        _agregar_carpeta(adjuntos_dir, "Adjuntos/Rendiciones")
        _agregar_carpeta(adjuntos_facturas_dir, "Adjuntos/Facturas")

        if code_dir.exists():
            for p in code_dir.rglob("*"):
                if not p.is_file():
                    continue
                partes = p.relative_to(code_dir).parts
                if any(parte in _EXCLUIR_RESPALDO_CODIGO for parte in partes):
                    continue
                zf.write(p, f"Codigo/{p.relative_to(code_dir)}")

        zf.writestr("LEEME.txt", _leeme_respaldo(fecha))
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Módulo 5 · Comparación con la cartola del banco
# ---------------------------------------------------------------------------

_FECHA_CARTOLA_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")


def _monto_cartola(campo: str) -> int:
    """Convierte un campo de monto de la cartola ('+0007000000', '00000000000')
    a entero en pesos. El signo indica si el campo tiene valor, no si es
    negativo (cargo/abono ya están en columnas separadas)."""
    campo = (campo or "").strip()
    if not campo:
        return 0
    digitos = campo[1:] if campo[0] in "+-" else campo
    digitos = digitos.strip() or "0"
    try:
        return int(digitos)
    except ValueError:
        return 0


def parsear_cartola_banco_chile(contenido: bytes) -> list[dict]:
    """Parsea un .txt de cartola del Banco de Chile ('CartolaEmitida...txt').

    Formato real (confirmado con un archivo de ejemplo): la primera línea trae
    el nombre/RUT/cuenta, la segunda es el encabezado de columnas, y cada línea
    de movimiento viene entre comillas con campos separados por ';':
        Fecha;Detalle Movimiento;Cheque o Cargo;Deposito o Abono;Saldo;
        Docto. Nro.;Trn;Caja;Sucursal
    Los montos vienen como signo + dígitos con ceros a la izquierda, sin
    decimales (pesos chilenos), p. ej. "+0007000000" = 7.000.000.

    Devuelve una lista de dicts: fecha (YYYY-MM-DD), detalle, flujo
    ('Ingreso' si hay abono, 'Egreso' si hay cargo), monto, canal. Las líneas
    que no son movimientos (encabezados) se descartan.
    """
    texto = None
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            texto = contenido.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if texto is None:
        texto = contenido.decode("utf-8", errors="replace")

    movimientos: list[dict] = []
    for linea in texto.splitlines():
        linea = linea.strip()
        if not linea:
            continue
        if linea.startswith('"') and linea.endswith('"'):
            linea = linea[1:-1]
        campos = linea.split(";")
        if len(campos) < 4:
            continue
        fecha_raw = campos[0].strip()
        if not _FECHA_CARTOLA_RE.match(fecha_raw):
            continue  # descarta encabezados y cualquier línea que no sea un movimiento
        detalle = campos[1].strip()
        cargo = _monto_cartola(campos[2])
        abono = _monto_cartola(campos[3]) if len(campos) > 3 else 0
        canal = campos[8].strip() if len(campos) > 8 else ""
        if abono > 0:
            flujo, monto = "Ingreso", abono
        elif cargo > 0:
            flujo, monto = "Egreso", cargo
        else:
            continue
        dd, mm, yyyy = fecha_raw.split("/")
        movimientos.append({
            "fecha": f"{yyyy}-{mm}-{dd}",
            "detalle": detalle,
            "flujo": flujo,
            "monto": monto,
            "canal": canal,
        })
    return movimientos


# Prefijos típicos del detalle de movimiento del banco que anteceden al
# nombre de la contraparte real (se descartan para poder comparar nombres).
_PREFIJOS_DETALLE_BANCO = (
    "APP-TRASPASO A:", "APP-TRASPASO DE:", "TRASPASO A:", "TRASPASO DE:",
)

# Palabras muy comunes en razones sociales/detalles que no sirven para
# distinguir una contraparte de otra (se ignoran al comparar nombres).
_STOPWORDS_CONTRAPARTE = {
    "spa", "sa", "ltda", "limitada", "eirl", "de", "del", "la", "el",
    "los", "las", "y", "a",
}


def _normalizar_texto(s: str) -> str:
    s = (s or "").lower()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9\s]", " ", s)


def _tokens_contraparte(s: str) -> set[str]:
    return {
        t for t in _normalizar_texto(s).split()
        if len(t) > 2 and t not in _STOPWORDS_CONTRAPARTE
    }


def _contraparte_app(descripcion: str) -> str:
    """La descripción de un movimiento de la app tiene forma
    "Tipo de documento N° folio · Razón social" (ver movimientos_en_rango);
    la contraparte es lo que va después del "·"."""
    descripcion = descripcion or ""
    if "·" in descripcion:
        return descripcion.split("·")[-1].strip()
    return descripcion


def _contraparte_banco(detalle: str) -> str:
    d = (detalle or "").strip()
    for prefijo in _PREFIJOS_DETALLE_BANCO:
        if d.upper().startswith(prefijo):
            return d[len(prefijo):].strip()
    return d


def comparar_cc(movs_app: list[dict], movs_banco: list, tolerancia_dias: int = 3) -> dict:
    """Compara los movimientos de la app (`movimientos_en_rango`) contra los de
    la cartola del banco (`cc_banco_en_rango`).

    Calce: mismo flujo (Ingreso/Egreso) y mismo monto exacto, con la fecha
    dentro de una tolerancia de `tolerancia_dias` días (las transferencias a
    veces se registran/abonan uno o dos días después). Es un calce voraz
    (greedy): a cada movimiento de la app se le busca, entre los movimientos
    de banco aún libres, el mejor candidato — priorizando primero los que
    además comparten alguna palabra del nombre de la contraparte (evita
    calces falsos cuando dos movimientos distintos comparten el mismo monto
    por coincidencia) y, entre esos, la fecha más cercana.

    Cada calce trae una "confianza":
    - "alta": el nombre de la contraparte coincide (en la app y en el banco).
    - "media": mismo monto y misma fecha exacta, pero no se pudo confirmar
      el nombre (p. ej. pagos por un agregador que no menciona a quién).
    - "revisar": mismo monto pero con fecha distinta Y sin coincidencia de
      nombre — el calce es posible pero podría ser una coincidencia; conviene
      confirmarlo a mano.

    Devuelve {"calzados": [...], "solo_app": [...], "solo_banco": [...]}.
    `calzados` trae pares {"app": <dict>, "banco": <dict>, "dif_dias": int,
    "confianza": str}.
    """
    from datetime import date as _date

    def _a_fecha(s: str) -> _date | None:
        try:
            return _date.fromisoformat(s)
        except (ValueError, TypeError):
            return None

    banco_libres = []
    for b in movs_banco:
        banco_libres.append({
            "fecha": b["fecha"], "detalle": b["detalle"],
            "flujo": b["flujo"], "monto": b["monto"], "canal": b["canal"],
        })

    calzados = []
    solo_app = []
    for m in movs_app:
        fecha_app = _a_fecha(m["fecha"])
        tokens_app = _tokens_contraparte(_contraparte_app(m["descripcion"]))
        candidatos = [
            b for b in banco_libres
            if b["flujo"] == m["flujo"] and b["monto"] == m["monto"]
        ]
        evaluados = []
        for b in candidatos:
            fecha_b = _a_fecha(b["fecha"])
            dif = 0 if fecha_app is None or fecha_b is None else abs((fecha_b - fecha_app).days)
            if dif > tolerancia_dias:
                continue
            tokens_b = _tokens_contraparte(_contraparte_banco(b["detalle"]))
            coincide_nombre = bool(tokens_app & tokens_b)
            evaluados.append((b, dif, coincide_nombre))
        # Prioriza: coincidencia de nombre primero, luego menor diferencia de días.
        evaluados.sort(key=lambda t: (not t[2], t[1]))
        if evaluados:
            mejor, mejor_dif, coincide_nombre = evaluados[0]
            if coincide_nombre:
                confianza = "alta"
            elif mejor_dif == 0:
                confianza = "media"
            else:
                confianza = "revisar"
            banco_libres.remove(mejor)
            calzados.append({
                "app": m, "banco": mejor, "dif_dias": mejor_dif, "confianza": confianza,
            })
        else:
            solo_app.append(m)

    calzados.sort(key=lambda c: c["app"]["fecha"] or "")
    solo_app.sort(key=lambda m: m["fecha"] or "")
    banco_libres.sort(key=lambda b: b["fecha"] or "")
    return {"calzados": calzados, "solo_app": solo_app, "solo_banco": banco_libres}


def construir_excel_comparacion(comp: dict, desde: str, hasta: str) -> bytes:
    """Arma el .xlsx de 'Exportar Comparación': tres hojas (Calzados, Solo
    banco, Solo app) a partir del resultado de `comparar_cc`."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    verde = "FF009406"
    tinta = "FF0A0A0A"
    rojo = "FFC0392B"
    ambar = "FFB8860B"

    wb = Workbook()

    def _titulo(ws, texto):
        ws["A1"] = texto
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=tinta)
        ws["A2"] = f"Rango: {desde} a {hasta}"
        ws["A2"].font = Font(name="Calibri", size=10, color="FF666666")

    def _encabezados(ws, textos, fila, color=tinta):
        borde = Border(bottom=Side(style="thin", color="FFDDDDDD"))
        for col, texto in enumerate(textos, start=1):
            c = ws.cell(row=fila, column=col, value=texto)
            c.font = Font(bold=True, color="FFFFFFFF")
            c.fill = PatternFill("solid", fgColor=color)
        return borde

    # --- Hoja 1: Calzados ---
    ws1 = wb.active
    ws1.title = "Calzados"
    n_revisar = sum(1 for c in comp["calzados"] if c["confianza"] == "revisar")
    titulo1 = f"Movimientos calzados ({len(comp['calzados'])})"
    if n_revisar:
        titulo1 += f" · {n_revisar} para revisar"
    _titulo(ws1, titulo1)
    encabezados1 = ["Fecha app", "Descripción app", "Flujo", "Monto",
                     "Fecha banco", "Detalle banco", "Dif. días", "Confianza"]
    borde = _encabezados(ws1, encabezados1, 4, verde)
    fila = 5
    etiqueta_confianza = {"alta": "Alta", "media": "Media", "revisar": "Revisar"}
    relleno_revisar = PatternFill("solid", fgColor="FFFCE8B8")
    for c in comp["calzados"]:
        a, b = c["app"], c["banco"]
        ws1.cell(row=fila, column=1, value=a["fecha"])
        ws1.cell(row=fila, column=2, value=a["descripcion"])
        cf = ws1.cell(row=fila, column=3, value=a["flujo"])
        cf.font = Font(bold=True, color=verde if a["flujo"] == "Ingreso" else rojo)
        cm = ws1.cell(row=fila, column=4, value=a["monto"])
        cm.number_format = '"$"#,##0'
        ws1.cell(row=fila, column=5, value=b["fecha"])
        ws1.cell(row=fila, column=6, value=b["detalle"])
        ws1.cell(row=fila, column=7, value=c["dif_dias"])
        cc = ws1.cell(row=fila, column=8, value=etiqueta_confianza[c["confianza"]])
        if c["confianza"] == "revisar":
            cc.font = Font(bold=True, color=ambar)
        for col in range(1, 9):
            celda = ws1.cell(row=fila, column=col)
            celda.border = borde
            if c["confianza"] == "revisar":
                celda.fill = relleno_revisar
        fila += 1
    anchos1 = {1: 13, 2: 46, 3: 11, 4: 14, 5: 13, 6: 46, 7: 11, 8: 12}
    for col, w in anchos1.items():
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = "A5"

    # --- Hoja 2: Solo banco (no registrado en la app) ---
    ws2 = wb.create_sheet("Solo banco")
    _titulo(ws2, f"Movimientos solo en el banco, sin registrar en la app ({len(comp['solo_banco'])})")
    borde = _encabezados(ws2, ["Fecha", "Detalle", "Flujo", "Monto", "Canal"], 4, ambar)
    fila = 5
    for b in comp["solo_banco"]:
        ws2.cell(row=fila, column=1, value=b["fecha"])
        ws2.cell(row=fila, column=2, value=b["detalle"])
        cf = ws2.cell(row=fila, column=3, value=b["flujo"])
        cf.font = Font(bold=True, color=verde if b["flujo"] == "Ingreso" else rojo)
        cm = ws2.cell(row=fila, column=4, value=b["monto"])
        cm.number_format = '"$"#,##0'
        ws2.cell(row=fila, column=5, value=b["canal"])
        for col in range(1, 6):
            ws2.cell(row=fila, column=col).border = borde
        fila += 1
    anchos2 = {1: 13, 2: 46, 3: 11, 4: 14, 5: 12}
    for col, w in anchos2.items():
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = "A5"

    # --- Hoja 3: Solo app (no aparece en el banco) ---
    ws3 = wb.create_sheet("Solo app")
    _titulo(ws3, f"Movimientos solo en la app, sin aparecer en la cartola ({len(comp['solo_app'])})")
    borde = _encabezados(ws3, ["Fecha", "Descripción", "Flujo", "Monto", "Origen"], 4, ambar)
    fila = 5
    for m in comp["solo_app"]:
        ws3.cell(row=fila, column=1, value=m["fecha"])
        ws3.cell(row=fila, column=2, value=m["descripcion"])
        cf = ws3.cell(row=fila, column=3, value=m["flujo"])
        cf.font = Font(bold=True, color=verde if m["flujo"] == "Ingreso" else rojo)
        cm = ws3.cell(row=fila, column=4, value=m["monto"])
        cm.number_format = '"$"#,##0'
        ws3.cell(row=fila, column=5, value=m["origen"])
        for col in range(1, 6):
            ws3.cell(row=fila, column=col).border = borde
        fila += 1
    anchos3 = {1: 13, 2: 46, 3: 11, 4: 14, 5: 12}
    for col, w in anchos3.items():
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
