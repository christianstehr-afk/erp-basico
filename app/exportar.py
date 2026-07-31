"""
Módulo 5 · Generación de archivos para el export contable.

- construir_excel: listado de movimientos (ingresos/egresos) a un .xlsx.
- construir_zip_rendiciones: un PDF por rendición (información + adjuntos, una
  imagen por página) empaquetados en un .zip.

Estas funciones no tocan la base de datos; reciben los datos ya consultados.
"""
from __future__ import annotations

import io
import os
import re
import zipfile
from pathlib import Path

from .db import codigo_rendicion

# Extensiones de imagen que se pintan "una por página" en el PDF.
IMG_EXT = {".png", ".jpg", ".jpeg", ".gif", ".bmp", ".tif", ".tiff", ".webp"}

# Logo E-Auto para la portada del PDF de rendición. Se descarga UNA vez y se
# cachea en disco (evita depender de internet en cada PDF); si la descarga
# falla, la portada se dibuja igual pero sin logo (nunca debe tumbar el PDF).
_LOGO_URL = "https://www.e-auto.global/assets/logos/eauto_logo_new.webp"
_LOGO_CACHE = Path(os.environ.get(
    "LOGO_CACHE_PATH", Path(__file__).resolve().parent.parent / "data" / "eauto_logo.png"
))


def _logo_reader():
    """Devuelve un ImageReader del logo E-Auto (cacheado), o None si no se pudo obtener."""
    from reportlab.lib.utils import ImageReader

    try:
        if not (_LOGO_CACHE.exists() and _LOGO_CACHE.stat().st_size > 0):
            import requests
            from PIL import Image

            resp = requests.get(_LOGO_URL, timeout=10)
            resp.raise_for_status()
            img = Image.open(io.BytesIO(resp.content)).convert("RGBA")
            _LOGO_CACHE.parent.mkdir(parents=True, exist_ok=True)
            img.save(_LOGO_CACHE, "PNG")
        return ImageReader(str(_LOGO_CACHE))
    except Exception:
        return None


def _miles(n: int) -> str:
    """Formato de miles con punto (estilo chileno): 1234567 -> 1.234.567."""
    return "{:,.0f}".format(n or 0).replace(",", ".")


def _nombre_archivo(nombre: str, sufijo: str) -> str:
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", (nombre or "").strip()) or "rendicion"
    return f"{base[:80]}{sufijo}"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def construir_excel(movimientos: list[dict], desde: str, hasta: str) -> bytes:
    """Arma el .xlsx del listado de movimientos y lo devuelve en bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    verde = "FF009406"
    tinta = "FF0A0A0A"
    rojo = "FFC0392B"
    gris = "FFF2F2F2"

    # Título y rango.
    ws["A1"] = "Movimientos de caja"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=tinta)
    ws["A2"] = f"Rango: {desde} a {hasta}"
    ws["A2"].font = Font(name="Calibri", size=10, color="FF666666")

    encabezados = ["Fecha", "Ingreso/Egreso", "Descripción", "Monto"]
    fila_head = 4
    borde = Border(bottom=Side(style="thin", color="FFDDDDDD"))
    for col, texto in enumerate(encabezados, start=1):
        c = ws.cell(row=fila_head, column=col, value=texto)
        c.font = Font(bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=tinta)
        c.alignment = Alignment(horizontal="right" if texto == "Monto" else "left")

    total_ing = 0
    total_egr = 0
    fila = fila_head + 1
    for m in movimientos:
        es_ing = m["flujo"] == "Ingreso"
        if es_ing:
            total_ing += m["monto"]
        else:
            total_egr += m["monto"]
        ws.cell(row=fila, column=1, value=m["fecha"])
        cflujo = ws.cell(row=fila, column=2, value=m["flujo"])
        cflujo.font = Font(bold=True, color=verde if es_ing else rojo)
        ws.cell(row=fila, column=3, value=m["descripcion"])
        cmonto = ws.cell(row=fila, column=4, value=m["monto"])
        cmonto.number_format = '"$"#,##0'
        cmonto.alignment = Alignment(horizontal="right")
        for col in range(1, 5):
            ws.cell(row=fila, column=col).border = borde
        fila += 1

    # Totales.
    fila += 1
    ws.cell(row=fila, column=3, value="Total ingresos").font = Font(bold=True, color=verde)
    ci = ws.cell(row=fila, column=4, value=total_ing)
    ci.number_format = '"$"#,##0'; ci.font = Font(bold=True, color=verde)
    fila += 1
    ws.cell(row=fila, column=3, value="Total egresos").font = Font(bold=True, color=rojo)
    ce = ws.cell(row=fila, column=4, value=total_egr)
    ce.number_format = '"$"#,##0'; ce.font = Font(bold=True, color=rojo)
    fila += 1
    ws.cell(row=fila, column=3, value="Neto (ingresos − egresos)").font = Font(bold=True, color=tinta)
    cn = ws.cell(row=fila, column=4, value=total_ing - total_egr)
    cn.number_format = '"$"#,##0'; cn.font = Font(bold=True, color=tinta)
    ws.cell(row=fila, column=3).fill = PatternFill("solid", fgColor=gris)
    ws.cell(row=fila, column=4).fill = PatternFill("solid", fgColor=gris)

    anchos = {1: 14, 2: 16, 3: 60, 4: 16}
    for col, w in anchos.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF por rendición
# ---------------------------------------------------------------------------

def _pagina_info(c, r, items, pagos) -> None:
    """Dibuja la página de información (portada) de una rendición."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm

    W, H = A4
    mx = 20 * mm
    y = H - 22 * mm

    # Logo: justificado a la derecha, a la misma altura del título (nombre),
    # sin ocupar espacio en el flujo vertical del texto de la izquierda.
    logo = _logo_reader()
    if logo is not None:
        try:
            iw, ih = logo.getSize()
            logo_h = 14 * mm
            logo_w = logo_h * iw / ih
            titulo_y = y - 9 * mm  # misma y donde se dibuja el título más abajo
            c.drawImage(logo, W - mx - logo_w, titulo_y - 3 * mm, logo_w, logo_h,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass  # un logo que no se pudo dibujar no debe tumbar el PDF

    c.setFillColorRGB(0, 0.58, 0.023)  # verde e-auto
    c.setFont("Helvetica-Bold", 9)
    c.drawString(mx, y, f"E-AUTO · RENDICIÓN {codigo_rendicion(r['id'])}")
    y -= 9 * mm
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(mx, y, (r["nombre"] or "")[:70])
    y -= 7 * mm
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.setFont("Helvetica", 10)
    c.drawString(mx, y, f"Fecha de rendición: {r['fecha']}")
    y -= 12 * mm

    total = r["total"] or 0
    pagado = r["pagado"] or 0
    saldo = total - pagado
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(mx, y, f"Total: ${_miles(total)}    Pagado: ${_miles(pagado)}    Saldo: ${_miles(saldo)}")
    y -= 12 * mm

    # Ítems
    c.setFont("Helvetica-Bold", 10)
    c.drawString(mx, y, "Ítems")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, y, "DESCRIPCIÓN")
    c.drawString(mx + 95 * mm, y, "N° DOC")
    c.drawRightString(W - mx, y, "MONTO")
    y -= 2 * mm
    c.setStrokeColorRGB(0.85, 0.85, 0.85)
    c.line(mx, y, W - mx, y)
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    for it in items:
        if y < 30 * mm:
            c.showPage()
            y = H - 22 * mm
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.1, 0.1, 0.1)
        c.drawString(mx, y, (it["descripcion"] or "")[:60])
        c.drawString(mx + 95 * mm, y, str(it["numero_doc"] or "—")[:18])
        c.drawRightString(W - mx, y, f"${_miles(it['monto'])}")
        y -= 5.5 * mm

    # Pagos
    y -= 6 * mm
    if y < 40 * mm:
        c.showPage()
        y = H - 22 * mm
    c.setFont("Helvetica-Bold", 10)
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.drawString(mx, y, "Pagos registrados")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, y, "FECHA")
    c.drawRightString(W - mx, y, "MONTO")
    y -= 2 * mm
    c.line(mx, y, W - mx, y)
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    for p in pagos:
        if y < 20 * mm:
            c.showPage()
            y = H - 22 * mm
            c.setFont("Helvetica", 9)
            c.setFillColorRGB(0.1, 0.1, 0.1)
        c.drawString(mx, y, str(p["fecha"]))
        c.drawRightString(W - mx, y, f"${_miles(p['monto'])}")
        y -= 5.5 * mm
    c.showPage()


def _pagina_imagen(c, path: str, caption: str) -> None:
    """Dibuja una imagen ocupando la página (con leyenda)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader

    W, H = A4
    mx = 15 * mm
    top = H - 15 * mm
    img = ImageReader(path)
    iw, ih = img.getSize()
    max_w = W - 2 * mx
    max_h = H - 30 * mm  # deja espacio para la leyenda
    escala = min(max_w / iw, max_h / ih)
    w = iw * escala
    h = ih * escala
    x = (W - w) / 2
    y = top - h
    c.drawImage(img, x, y, w, h, preserveAspectRatio=True, anchor="n")
    c.setFont("Helvetica", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, 10 * mm, f"Adjunto: {caption}"[:110])
    c.showPage()


def _pdf_de_rendicion(r, items, adjuntos, pagos) -> bytes:
    """Devuelve el PDF (bytes) de una rendición: info + imágenes + PDFs adjuntos."""
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from pypdf import PdfReader, PdfWriter

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    _pagina_info(c, r, items, pagos)
    # Imágenes: una por página.
    for a in adjuntos:
        ext = Path(a["nombre_archivo"] or "").suffix.lower()
        if ext in IMG_EXT and Path(a["path"]).exists():
            try:
                _pagina_imagen(c, a["path"], a["nombre_archivo"])
            except Exception:
                pass
    c.save()

    writer = PdfWriter()
    for page in PdfReader(io.BytesIO(buf.getvalue())).pages:
        writer.add_page(page)
    # PDFs adjuntos: se anexan al final tal cual.
    for a in adjuntos:
        ext = Path(a["nombre_archivo"] or "").suffix.lower()
        if ext == ".pdf" and Path(a["path"]).exists():
            try:
                for page in PdfReader(a["path"]).pages:
                    writer.add_page(page)
            except Exception:
                pass
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


def _pdf_de_gestion_pago(f, pagos, cfg, incluye_original: bool = False,
                         incluye_adjuntos: bool = False) -> bytes:
    """Devuelve el PDF (bytes) del detalle de gestión de una factura (pago a
    proveedores / cobro a clientes): resumen + movimientos parciales.

    Si `incluye_original` es True, deja una nota indicando que el documento
    original de la factura viene a continuación (el llamador es quien anexa
    esas páginas, con `anexar_pdf`, porque requiere una sesión SII activa).
    Si `incluye_adjuntos` es True, deja una nota equivalente para los
    adjuntos de la gestión (el llamador los anexa con `anexar_archivos`)."""
    import textwrap

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas

    from .db import codigo_rendicion

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    W, H = A4
    mx = 20 * mm
    y = H - 22 * mm

    logo = _logo_reader()
    if logo is not None:
        try:
            iw, ih = logo.getSize()
            logo_h = 14 * mm
            logo_w = logo_h * iw / ih
            titulo_y = y - 9 * mm
            c.drawImage(logo, W - mx - logo_w, titulo_y - 3 * mm, logo_w, logo_h,
                        preserveAspectRatio=True, mask="auto")
        except Exception:
            pass

    rechazada = bool(f["fecha_reclamo"])
    total = f["total"] or 0
    pagado = f["pagado"] or 0
    saldo = total - pagado
    pagada = pagado >= total and total > 0
    estado = "Rechazada" if rechazada else (cfg["estado_ok"] if pagada else "Pendiente")

    c.setFillColorRGB(0, 0.58, 0.023)  # verde e-auto
    c.setFont("Helvetica-Bold", 9)
    c.drawString(mx, y, f"E-AUTO · {cfg['titulo'].upper()}")
    y -= 9 * mm
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(mx, y, f"{f['documento']} · Folio {f['folio']}"[:80])
    y -= 7 * mm
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.setFont("Helvetica", 10)
    c.drawString(mx, y, f"{f['razon_social'] or ''} · {f['rut_contraparte'] or ''}"[:100])
    y -= 6 * mm
    c.drawString(mx, y, f"Fecha de emisión: {f['fecha_emision'] or '—'}")
    y -= 12 * mm

    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 11)
    c.drawString(
        mx, y,
        f"Monto factura: ${_miles(total)}    {cfg['label_pagado']}: ${_miles(pagado)}    "
        f"Saldo: ${_miles(saldo)}",
    )
    y -= 7 * mm
    c.setFont("Helvetica-Bold", 10)
    if rechazada:
        c.setFillColorRGB(0.75, 0.15, 0.15)
    elif pagada:
        c.setFillColorRGB(0, 0.58, 0.023)
    else:
        c.setFillColorRGB(0.05, 0.4, 0.85)
    c.drawString(mx, y, f"Estado: {estado}")
    y -= 10 * mm

    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(mx, y, f"Fecha de {cfg['accion']} tope: {f['fecha_pago_tope'] or '—'}")
    y -= 10 * mm

    if f["descripcion"]:
        c.setFont("Helvetica-Bold", 10)
        c.drawString(mx, y, "Descripción")
        y -= 6 * mm
        c.setFont("Helvetica", 9)
        c.setFillColorRGB(0.15, 0.15, 0.15)
        for linea in textwrap.wrap(f["descripcion"], 105) or [""]:
            if y < 25 * mm:
                c.showPage()
                y = H - 22 * mm
                c.setFont("Helvetica", 9)
                c.setFillColorRGB(0.15, 0.15, 0.15)
            c.drawString(mx, y, linea)
            y -= 5 * mm
        y -= 6 * mm

    # Movimientos parciales
    if y < 40 * mm:
        c.showPage()
        y = H - 22 * mm
    c.setFillColorRGB(0.04, 0.04, 0.04)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(mx, y, f"{cfg['accion'].capitalize()}s parciales")
    y -= 6 * mm
    c.setFont("Helvetica-Bold", 8)
    c.setFillColorRGB(0.4, 0.4, 0.4)
    c.drawString(mx, y, "FECHA")
    c.drawString(mx + 40 * mm, y, "VÍA")
    c.drawRightString(W - mx, y, "MONTO")
    y -= 2 * mm
    c.setStrokeColorRGB(0.85, 0.85, 0.85)
    c.line(mx, y, W - mx, y)
    y -= 5 * mm
    c.setFont("Helvetica", 9)
    c.setFillColorRGB(0.1, 0.1, 0.1)
    if pagos:
        for p in pagos:
            if y < 20 * mm:
                c.showPage()
                y = H - 22 * mm
                c.setFont("Helvetica", 9)
                c.setFillColorRGB(0.1, 0.1, 0.1)
            if p["rendicion_id"]:
                via = f"Rendición {codigo_rendicion(p['rendicion_id'])}"
            elif p["externo"]:
                via = "Externo"
            else:
                via = "—"
            c.drawString(mx, y, str(p["fecha"]))
            c.drawString(mx + 40 * mm, y, via[:38])
            c.drawRightString(W - mx, y, f"${_miles(p['monto'])}")
            y -= 5.5 * mm
    else:
        c.setFillColorRGB(0.5, 0.5, 0.5)
        c.drawString(mx, y, f"Aún no hay {cfg['accion']}s registrados.")
        y -= 5.5 * mm

    if incluye_original:
        y -= 8 * mm
        if y < 20 * mm:
            c.showPage()
            y = H - 22 * mm
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(mx, y, "— Documento original de la factura a continuación —")

    if incluye_adjuntos:
        y -= 6 * mm
        if y < 20 * mm:
            c.showPage()
            y = H - 22 * mm
        c.setFont("Helvetica-Oblique", 9)
        c.setFillColorRGB(0.4, 0.4, 0.4)
        c.drawString(mx, y, "— Adjuntos de la gestión a continuación —")

    c.showPage()
    c.save()
    return buf.getvalue()


def anexar_archivos(base: bytes, adjuntos: list[dict]) -> bytes:
    """Anexa al final de un PDF los archivos adjuntos de la gestión de una
    factura (ver `factura_adjuntos` en db.py): imágenes se dibujan una por
    página (con leyenda con el nombre del archivo), PDFs se agregan tal
    cual. Un adjunto que no existe en disco o no se puede leer se salta sin
    interrumpir a los demás — nunca debe tumbar la generación del PDF.
    """
    if not adjuntos:
        return base

    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas

    resultado = base

    imagenes = [
        a for a in adjuntos
        if Path(a["nombre_archivo"] or "").suffix.lower() in IMG_EXT
        and a["path"] and Path(a["path"]).exists()
    ]
    if imagenes:
        buf = io.BytesIO()
        c = canvas.Canvas(buf, pagesize=A4)
        for a in imagenes:
            try:
                _pagina_imagen(c, a["path"], a["nombre_archivo"])
            except Exception:
                pass  # una imagen que no se pudo dibujar no debe tumbar el resto
        c.save()
        resultado = anexar_pdf(resultado, buf.getvalue())

    for a in adjuntos:
        ext = Path(a["nombre_archivo"] or "").suffix.lower()
        if ext == ".pdf" and a["path"] and Path(a["path"]).exists():
            try:
                resultado = anexar_pdf(resultado, Path(a["path"]).read_bytes())
            except Exception:
                pass  # un PDF adjunto ilegible no debe tumbar el resto

    return resultado


def anexar_pdf(base: bytes, extra: bytes | None) -> bytes:
    """Agrega las páginas de `extra` al final de `base`.

    Si `extra` es None o no se puede leer como PDF, devuelve `base` sin
    cambios (nunca debe tumbar la generación del PDF de gestión).
    """
    if not extra:
        return base
    from pypdf import PdfReader, PdfWriter

    try:
        writer = PdfWriter()
        for page in PdfReader(io.BytesIO(base)).pages:
            writer.add_page(page)
        for page in PdfReader(io.BytesIO(extra)).pages:
            writer.add_page(page)
        out = io.BytesIO()
        writer.write(out)
        return out.getvalue()
    except Exception:
        return base


def construir_zip_rendiciones(rendiciones: list[dict]) -> bytes:
    """Empaqueta un PDF por rendición en un .zip.

    `rendiciones` es una lista de dicts con: r (fila), items, adjuntos, pagos.
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        usados: set[str] = set()
        for rd in rendiciones:
            r = rd["r"]
            pdf = _pdf_de_rendicion(r, rd["items"], rd["adjuntos"], rd["pagos"])
            cod = codigo_rendicion(r["id"])
            nombre = f"{cod}_" + _nombre_archivo(r["nombre"], ".pdf")
            while nombre in usados:
                nombre = f"{cod}_" + _nombre_archivo(r["nombre"], "_x.pdf")
            usados.add(nombre)
            zf.writestr(nombre, pdf)
    return buf.getvalue()
